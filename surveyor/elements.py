# -*- coding: utf-8 -*-


from __future__ import unicode_literals, absolute_import

import abc
import collections
import sys

import openpyxl
import openpyxl.utils
import openpyxl.comments
import openpyxl.styles
import six

import surveyor.classes.simple
import surveyor.exceptions
import surveyor.utils

# noinspection PyUnresolvedReferences
from six.moves import range, zip


DEFAULT_CLASSES_MODULE = "surveyor.classes.simple"
"""Default importable module with style classes."""


@six.add_metaclass(abc.ABCMeta)
class BaseElement(object):

    TAG_NAME = None

    DEFAULT_STYLER = None

    @property
    def classes(self):
        return self.parent.classes

    def __init__(self, element):
        if element.tag.lower() != self.TAG_NAME:
            raise surveyor.exceptions.UnexpectedTagError(element.tag, self.TAG_NAME)

        super(BaseElement, self).__init__()

        self.parent = None
        self.klass = None
        self.children = []

    def add(self, element):
        element.parent = self
        self.children.append(element)

    def process(self, element=None, *args, **kwargs):
        openpyxl_elements = self.collect(element, *args, **kwargs)

        self.apply_inline_styles(openpyxl_elements)
        self.stylize(openpyxl_elements)

        return openpyxl_elements

    @abc.abstractmethod
    def collect(self, element, *args, **kwargs):
        raise NotImplementedError("You have to define this method to invoke")

    def apply_inline_styles(self, elements):
        return elements

    def stylize(self, elements):
        return elements

    def get_class(self):
        if self.klass is None:
            return self.DEFAULT_STYLER

        class_obj = getattr(self.classes, self.klass, None)
        if class_obj is None:
            raise surveyor.exceptions.CannotFindElementClass(self.classes, self.klass)

        return class_obj


class WorkBook(BaseElement):

    TAG_NAME = "workbook"

    ATTR_CLASSES = "classes"

    @property
    def classes(self):
        return self.class_module

    def __init__(self, element):
        super(WorkBook, self).__init__(element)

        module_name = self.make_module_name(element)

        try:
            self.class_module = __import__(module_name)
        except Exception as exc:
            raise surveyor.exceptions.CannotImportClassError(module_name, exc)
        else:
            self.class_module = sys.modules[module_name]

    def add(self, element):
        super(WorkBook, self).add(element)

        if not element.name:
            element.name = "Sheet{0}".format(len(self.children))

    def collect(self, element, *args, **kwargs):
        book = openpyxl.Workbook(encoding="utf-8", guess_types=True)
        book.worksheets = []

        for sheet in self.children:
            sheet_element = book.create_sheet(title=sheet.name)
            sheet.process(sheet_element)

        return book

    def make_module_name(self, element):
        return element.attrib.get(self.ATTR_CLASSES, DEFAULT_CLASSES_MODULE)


class Sheet(BaseElement):

    TAG_NAME = "sheet"

    ATTR_AUTOSIZE = "autosize"
    ATTR_NAME = "name"
    ATTR_CLASS = "class"
    ATTR_FREEZE_ROW = "freeze-row"
    ATTR_FREEZE_COLUMN = "freeze-column"

    DEFAULT_WIDTH = 10
    WIDTH_ADDITION = 1

    DEFAULT_STYLER = surveyor.classes.simple.Sheet

    @staticmethod
    def apply_autosize(sheet, default_width=DEFAULT_WIDTH, width_addition=WIDTH_ADDITION):
        column_width = collections.defaultdict(lambda: default_width)

        for row in range(1, sheet.max_row):
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                if not value:
                    continue
                if not isinstance(value, six.string_types):
                    value = six.text_type(value)
                if value.startswith("="):  # Skip formulas
                    continue

                column_width[col] = max(column_width[col], len(value))

        for col in range(1, sheet.max_column + 1):
            try:
                column = sheet.column_dimensions[openpyxl.utils.get_column_letter(col)]
            except KeyError:
                pass
            else:
                column.auto_size = True
                column.width = column_width[col] + width_addition

    def __init__(self, element):
        super(Sheet, self).__init__(element)

        self.autosize = surveyor.utils.strtobool(element.attrib.get(self.ATTR_AUTOSIZE))
        self.klass = element.attrib.get(self.ATTR_CLASS)
        self.name = element.attrib.get(self.ATTR_NAME)
        self.freeze_row = element.attrib.get(self.ATTR_FREEZE_ROW)
        self.freeze_col = element.attrib.get(self.ATTR_FREEZE_COLUMN)

    def collect(self, element, *args, **kwargs):
        for table in self.children:
            table.process(element)

        return element

    def apply_inline_styles(self, element):
        if self.autosize:
            self.apply_autosize(element)

        freeze_row, freeze_col = self.freeze_row, self.freeze_col
        if freeze_row is not None or freeze_col is not None:
            if not freeze_col:
                freeze_col = 1
            if not freeze_row:
                freeze_row = 1
            element.freeze_panes = element.cell(row=int(freeze_row), column=int(freeze_col))

    def stylize(self, element):
        styler = self.get_class()(element)
        styler.stylize()

        return element


class Table(BaseElement):

    TAG_NAME = "table"

    ATTR_START_CELL = "startcell"
    ATTR_START_ROW = "startrow"
    ATTR_START_COLUMN = "startcolumn"
    ATTR_CLASS = "class"

    DEFAULT_START_CELL = "A1"

    DEFAULT_STYLER = surveyor.classes.simple.Table

    @classmethod
    def get_start_cell(cls, attrs):
        start_row = attrs[cls.ATTR_START_ROW]
        start_column = attrs[cls.ATTR_START_COLUMN]

        if start_column.isdigit():
            start_column = openpyxl.utils.get_column_letter(int(start_column))
        start_column = start_column.upper()

        return start_column + start_row

    def __init__(self, element):
        super(Table, self).__init__(element)

        self.klass = element.attrib.get(self.ATTR_CLASS)
        if self.ATTR_START_CELL in element.attrib:
            self.start_cell = element.attrib[self.ATTR_START_CELL]
        elif self.ATTR_START_ROW in element.attrib and self.ATTR_START_COLUMN in element.attrib:
            self.start_cell = self.get_start_cell(element.attrib)
        else:
            self.start_cell = self.DEFAULT_START_CELL

    def collect(self, element, *args, **kwargs):
        top_row, bottom_row, left_column, right_column = self.get_dimensions()

        for row, row_idx in zip(self.children, range(top_row, bottom_row)):
            row.process(element, row_idx, left_column, right_column)

        return element

    def stylize(self, element):
        top_row, bottom_row, left_column, right_column = self.get_dimensions()
        styler = self.get_class()(element, top_row, bottom_row, left_column, right_column)

        styler.stylize()

        return element

    def get_dimensions(self):
        left_column, top_row = openpyxl.utils.coordinate_from_string(self.start_cell)
        left_column = openpyxl.utils.column_index_from_string(left_column)

        bottom_row = top_row + len(self.children)
        right_column = left_column + max(len(row.children) for row in self.children)

        return top_row, bottom_row, left_column, right_column


class Row(BaseElement):

    TAG_NAME = "tr"

    ATTR_CLASS = "class"

    DEFAULT_STYLER = surveyor.classes.simple.Row

    def __init__(self, element):
        super(Row, self).__init__(element)

        self.klass = element.attrib.get(self.ATTR_CLASS)

    def collect(self, element, row_idx=1, left_column=1, right_column=1):
        cells = []

        for col_idx, cell in enumerate(self.children, start=left_column):
            cell.process(element, row_idx, col_idx)
            cells.append(cells)

        return cells

    def stylize(self, cells):
        styler = self.get_class()(cells)
        styler.stylize()

        return cells


StyleInfo = collections.namedtuple("StyleInfo", ["attribute", "constructor"])


class Cell(BaseElement):

    TAG_NAME = "td"

    ATTR_SEPARATOR = "-"

    ATTR_CLASS = "class"
    ATTR_NUMBER_FORMAT = "number_format"
    ATTR_HYPERLINK = "hyperlink"
    ATTR_COMMENT = "comment"
    ATTR_COMMENT_AUTHOR = "comment{0}author".format(ATTR_SEPARATOR)

    STYLE_ATTRIBUTES = {
        "font": StyleInfo("font", openpyxl.styles.Font),
        "pattern-fill": StyleInfo("fill", openpyxl.styles.PatternFill),
        "gradient-fill": StyleInfo("fill", openpyxl.styles.GradientFill),
        "alignment": StyleInfo("alignment", openpyxl.styles.Alignment),
        "protection": StyleInfo("protection", openpyxl.styles.Protection),
        # TODO border
    }
    STYLE_PREFIXES = tuple(STYLE_ATTRIBUTES.keys())

    DEFAULT_STYLER = surveyor.classes.simple.Cell

    def __init__(self, element):
        super(Cell, self).__init__(element)

        self.klass = element.attrib.get(self.ATTR_CLASS)
        self.value = surveyor.utils.guess_text(element.text)
        # check openpyxl.styles.numbers
        self.number_format = element.attrib.get(self.ATTR_NUMBER_FORMAT)
        self.hyperlink = element.attrib.get(self.ATTR_HYPERLINK)

        comment_text = element.attrib.get(self.ATTR_COMMENT)
        if comment_text:
            self.comment = openpyxl.comments.Comment(comment_text, element.attrib.get(self.ATTR_COMMENT_AUTHOR))
        else:
            self.comment = None

        self.styles_by_prefix = collections.defaultdict(dict)
        for attr, value in element.attrib.items():
            if attr.startswith(self.STYLE_PREFIXES):
                prefix, name = attr.rsplit(self.ATTR_SEPARATOR, 1)
                self.styles_by_prefix[prefix][name] = value
        # TODO check for multiple type of fills (maybe by the same StyleInfo.attribute)

    def collect(self, element, row_idx=1, col_idx=1):
        cell = element.cell(row=row_idx, column=col_idx)
        cell.value = self.value
        if self.hyperlink is not None:
            cell.hyperlink = self.hyperlink
        if self.comment is not None:
            cell.comment = self.comment

        return cell

    def apply_inline_styles(self, cell):
        if self.number_format is not None:
            cell.number_format = self.number_format

        for prefix, kwargs in self.styles_by_prefix.items():
            style_info = self.STYLE_ATTRIBUTES[prefix]
            style_object = style_info.constructor(**kwargs)
            setattr(cell, style_info.attribute, style_object)
        return cell

    def stylize(self, cell):
        styler = self.get_class()(cell)
        styler.stylize()

        return cell
