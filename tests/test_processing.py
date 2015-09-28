# -*- coding: utf-8 -*-


from __future__ import unicode_literals

import pytest
import openpyxl.styles

import surveyor.parse as parse


def test_check_sheet_names():
    xml = """
    <workbook>
        <sheet name="Worksheet1" />
        <sheet autosize="true" />
        <sheet />
        <sheet name="Custom" />
    </workbook>
    """.strip()

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    assert workbook.worksheets[0].title == "Worksheet1"
    assert workbook.worksheets[1].title == "Sheet2"
    assert workbook.worksheets[2].title == "Sheet3"
    assert workbook.worksheets[3].title == "Custom"


def test_fill_table():
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td>Name</td>
                    <td>Value</td>
                </tr>
                <tr>
                    <td>1</td>
                    <td>2text</td>
                </tr>
                <tr>
                    <td>3</td>
                    <td>4.0</td>
                </tr>
            </table>
            <table startcell="C10">
                <tr>
                    <td>Book</td>
                    <td>Title</td>
                </tr>
                <tr>
                    <td>111</td>
                    <td>222text</td>
                </tr>
                <tr>
                    <td>333</td>
                    <td>444.0</td>
                </tr>
            </table>
        </sheet>
        <sheet>
            <table>
                <tr>
                    <td>Book</td>
                    <td>Title</td>
                </tr>
                <tr>
                    <td>111</td>
                    <td>222text</td>
                </tr>
                <tr>
                    <td>333</td>
                    <td>444.0</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """.strip()

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    assert workbook.worksheets[0].cell(row=1, column=1).value == "Name"
    assert workbook.worksheets[0].cell(row=1, column=2).value == "Value"
    assert workbook.worksheets[0].cell(row=2, column=1).value == 1
    assert workbook.worksheets[0].cell(row=2, column=2).value == "2text"
    assert workbook.worksheets[0].cell(row=3, column=1).value == 3
    assert workbook.worksheets[0].cell(row=3, column=2).value == 4.0

    assert workbook.worksheets[0].cell(row=10, column=3).value == "Book"
    assert workbook.worksheets[0].cell(row=10, column=4).value == "Title"
    assert workbook.worksheets[0].cell(row=11, column=3).value == 111
    assert workbook.worksheets[0].cell(row=11, column=4).value == "222text"
    assert workbook.worksheets[0].cell(row=12, column=3).value == 333
    assert workbook.worksheets[0].cell(row=12, column=4).value == 444.0

    assert workbook.worksheets[1].cell(row=1, column=1).value == "Book"
    assert workbook.worksheets[1].cell(row=1, column=2).value == "Title"
    assert workbook.worksheets[1].cell(row=2, column=1).value == 111
    assert workbook.worksheets[1].cell(row=2, column=2).value == "222text"
    assert workbook.worksheets[1].cell(row=3, column=1).value == 333
    assert workbook.worksheets[1].cell(row=3, column=2).value == 444.0


# noinspection PyUnresolvedReferences
@pytest.mark.parametrize("number_format", (
    "0",
    "0.00",
    "#,##0",
    "#,##0.00"
))
def test_check_inline_number_format(number_format):
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td number_format="{0}">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """.format(number_format).strip()

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    assert workbook.worksheets[0].cell(row=1, column=1).number_format == number_format


# noinspection PyUnresolvedReferences
@pytest.mark.parametrize("hyperlink", (
    "http:",
    "aaa",
    "http://google.com",
))
def test_hyperlink(hyperlink):
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td hyperlink="{0}">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """.format(hyperlink).strip()

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    assert workbook.worksheets[0].cell(row=1, column=1).hyperlink == hyperlink


# noinspection PyUnresolvedReferences
@pytest.mark.parametrize("comment, author", (
    ("", None),
    ("comment", None),
    ("comment", "author"),
))
def test_comment(comment, author):
    attributes = "comment='{0}'".format(comment)
    if author is not None:
        attributes += " comment-author='{0}'".format(author)

    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td {0}>1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """.format(attributes).strip()

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    cell = workbook.worksheets[0].cell(row=1, column=1)
    if not comment:
        assert cell.comment is None
    else:
        assert cell.comment.text == comment
        if author:
            assert cell.comment.author == author


def test_font_styles():
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td font-name="name" font-sz="21" font-family="4">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    cell = workbook.worksheets[0].cell(row=1, column=1)
    assert cell.font.name == "name"
    assert cell.font.sz == 21
    assert cell.font.family == 4


def test_pattern_fill_styles():
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td pattern-fill-patternType="solid" pattern-fill-fgColor="0d0d0d">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    cell = workbook.worksheets[0].cell(row=1, column=1)
    assert cell.fill.patternType == "solid"
    assert cell.fill.fgColor == openpyxl.styles.Color("0d0d0d")


def test_gradient_fill_styles():
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td gradient-fill-degree="10" gradient-fill-left="5"
                        gradient-fill-stop="00ff00, ff0000">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    cell = workbook.worksheets[0].cell(row=1, column=1)
    assert cell.fill.degree == 10
    assert cell.fill.left == 5
    assert len(cell.fill.stop) == 2
    assert cell.fill.stop[0] == openpyxl.styles.Color("00ff00")
    assert cell.fill.stop[1] == openpyxl.styles.Color("ff0000")


def test_alignment_styles():
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td alignment-horizontal="right" alignment-vertical="top">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    cell = workbook.worksheets[0].cell(row=1, column=1)
    assert cell.alignment.horizontal == "right"
    assert cell.alignment.vertical == "top"


def test_protection_styles():
    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td protection-locked="0" protection-hidden="1">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    cell = workbook.worksheets[0].cell(row=1, column=1)
    assert not cell.protection.locked
    assert cell.protection.hidden


# noinspection PyUnresolvedReferences
@pytest.mark.parametrize("row, column", (
    (None, None),
    (1, None),
    (2, None),
    (None, 1),
    (None, 2),
    (2, 1),
    (2, 2),
    (1, 2),
))
def test_freeze_panes(row, column):
    freeze_attrs = ""
    if row:
        freeze_attrs += " freeze-row='{0}'".format(row)
    if column:
        freeze_attrs += " freeze-column='{0}'".format(column)
    xml = """
    <workbook>
        <sheet {0}>
            <table>
                <tr>
                    <td>1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """.format(freeze_attrs)

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    sheet = workbook.worksheets[0]
    freeze_panes = sheet.freeze_panes

    if row in (1, None) and column in (1, None):
        assert freeze_panes is None
    else:
        if row is None:
            row = 1
        if column is None:
            column = 1
        assert freeze_panes == sheet.cell(row=row, column=column).coordinate


# noinspection PyUnresolvedReferences
@pytest.mark.parametrize("border, text, style, color", (
    ("top", "thick", "thick", None),
    ("top", "thick, ff0000", "thick", "ff0000"),
    ("bottom", "thin,      00ff00", "thin", "00ff00"),
))
def test_border(border, text, style, color):

    xml = """
    <workbook>
        <sheet>
            <table>
                <tr>
                    <td border-{0}="{1}">1</td>
                </tr>
            </table>
        </sheet>
    </workbook>
    """.format(border, text)

    workbook = parse.parse_fileobj(xml)
    workbook = workbook.process()

    cell = workbook.worksheets[0].cell(row=1, column=1)
    cell_border = getattr(cell.border, border)
    assert cell_border.style == style
    if color is None:
        assert cell_border.color is None
    else:
        assert cell_border.color == openpyxl.styles.Color(color)
